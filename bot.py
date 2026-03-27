#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import tempfile
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from dotenv import load_dotenv
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN", "")
ADMIN_IDS = [int(id) for id in os.getenv("ADMIN_IDS", "").split(",") if id]

DATA_DIR = "data"
OUTPUT_DIR = "output"
TEMPLATES_DIR = "templates"
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(TEMPLATES_DIR, exist_ok=True)

# Ключевые слова для определения дохода
INCOME_KEYWORDS = [
    "оплата за товар", "оплата по договору", "оплата за услуги",
    "интернет решения", "озон", "по реестру", "оплата по контракту",
    "платеж по ден.треб", "за товар", "оплата за ооо", "оплата за ип",
    "оплата за ооо \"интернет решения\"", "оплата за ооо \"интернет решения\"",
]

EXCLUDE_KEYWORDS = [
    "собственных средств", "перевод собственных", "вывод собственных",
    "комиссия", "уплата налога", "страховые взносы"
]

IP_INN = "632312967829"
IP_FIO = "Леонтьев Артём Владиславович"
IP_OKTMO = "36701320"
IP_OKVED = "47.91"
IP_PHONE = ""

user_sessions = {}


class UserSession:
    def __init__(self, user_id):
        self.user_id = user_id
        self.bank_operations = []
        self.ens_data = {
            'insurance_accrued': 0,
            'insurance_paid': 0,
            'insurance_paid_dates': [],
            'penalties': 0
        }

    def add_bank_operations(self, operations):
        self.bank_operations.extend(operations)

    def set_ens_data(self, data):
        self.ens_data = data

    def reset(self):
        self.bank_operations = []
        self.ens_data = {
            'insurance_accrued': 0,
            'insurance_paid': 0,
            'insurance_paid_dates': [],
            'penalties': 0
        }


# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========

def safe_float(value):
    try:
        if pd.isna(value):
            return 0.0
        if isinstance(value, str):
            cleaned = value.replace(" ", "").replace(",", ".")
            return float(cleaned)
        return float(value)
    except:
        return 0.0


def parse_date(date_str):
    if isinstance(date_str, datetime):
        return date_str
    if isinstance(date_str, pd.Timestamp):
        return date_str.to_pydatetime()
    if isinstance(date_str, str):
        formats = ["%d.%m.%Y", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M"]
        for fmt in formats:
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except:
                continue
    return None


def is_income(purpose):
    purpose_lower = str(purpose).lower()
    for word in EXCLUDE_KEYWORDS:
        if word in purpose_lower:
            return False
    for word in INCOME_KEYWORDS:
        if word in purpose_lower:
            return True
    return False


def format_currency(amount):
    if amount == int(amount):
        return int(amount)
    return round(amount, 2)


def safe_write(ws, row, col, value):
    """Безопасная запись в ячейку с учетом объединенных ячеек"""
    try:
        cell = ws.cell(row=row, column=col)
        cell.value = value
    except AttributeError:
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and \
               merged_range.min_col <= col <= merged_range.max_col:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
        raise


# ========== УНИВЕРСАЛЬНЫЙ ПАРСИНГ ВЫПИСКИ ==========

def parse_bank_statement(file_path):
    """Универсальный парсинг Excel-выписки из любого банка"""
    try:
        # Читаем весь файл
        df_raw = pd.read_excel(file_path, header=None)
        
        # Ищем строку с данными (не заголовок)
        # Пропускаем первые строки до тех пор, пока не найдем строку с датой
        data_start_row = None
        for idx, row in df_raw.iterrows():
            first_cell = str(row.iloc[0]) if len(row) > 0 else ""
            # Ищем строку, где первая ячейка похожа на дату (число или формат дд.мм.гггг)
            if first_cell and (first_cell.replace('.', '').isdigit() or '.' in first_cell):
                if len(first_cell) >= 8 and '.' in first_cell:
                    data_start_row = idx
                    break
        
        if data_start_row is None:
            raise Exception("Не удалось найти строки с данными")
        
        # Извлекаем данные
        df_data = df_raw.iloc[data_start_row:].reset_index(drop=True)
        
        # Определяем колонки по первым строкам данных
        # Ищем колонки с датой, суммой и назначением
        col_date = None
        col_credit = None
        col_debit = None
        col_purpose = None
        
        # Проверяем первые 3 строки данных для определения колонок
        for col_idx in range(len(df_data.columns)):
            # Проверяем значения в первых строках
            sample_values = []
            for row_idx in range(min(5, len(df_data))):
                val = df_data.iloc[row_idx, col_idx]
                if pd.notna(val):
                    sample_values.append(str(val).lower())
            
            if not sample_values:
                continue
            
            # Определяем тип колонки
            sample_str = ' '.join(sample_values)
            
            # Дата
            if any('.' in v and len(v) >= 8 and v.replace('.', '').isdigit() for v in sample_values):
                if col_date is None:
                    col_date = col_idx
            
            # Кредит (доход) - ищем числа > 0
            elif any(v.replace('.', '').replace(',', '').replace('-', '').isdigit() and 
                     len(v) > 0 and v != '0' and not any(c.isalpha() for c in v) for v in sample_values):
                # Проверяем, что это положительное число
                try:
                    num = safe_float(sample_values[0])
                    if num > 0 and col_credit is None:
                        col_credit = col_idx
                except:
                    pass
            
            # Назначение платежа
            elif any(len(v) > 20 and ('оплата' in v or 'перевод' in v or 'платеж' in v) for v in sample_values):
                col_purpose = col_idx
        
        # Если не нашли кредит, ищем дебет (отрицательные суммы)
        if col_credit is None:
            for col_idx in range(len(df_data.columns)):
                val = df_data.iloc[0, col_idx]
                if pd.notna(val):
                    try:
                        num = safe_float(val)
                        if num < 0:
                            col_debit = col_idx
                            break
                    except:
                        pass
        
        if col_date is None:
            raise Exception("Не удалось определить колонку с датой")
        
        if col_credit is None and col_debit is None:
            raise Exception("Не удалось определить колонку с суммой")
        
        operations = []
        
        for idx, row in df_data.iterrows():
            try:
                date_val = row.iloc[col_date]
                if pd.isna(date_val):
                    continue
                date = parse_date(date_val)
                if not date:
                    continue
                
                amount = 0
                if col_credit is not None:
                    amount = safe_float(row.iloc[col_credit])
                elif col_debit is not None:
                    amount = -safe_float(row.iloc[col_debit])
                
                if amount == 0:
                    continue
                
                purpose = ""
                if col_purpose is not None:
                    purpose_val = row.iloc[col_purpose]
                    if pd.notna(purpose_val):
                        purpose = str(purpose_val)
                
                doc_num = f"п/п {idx+1}"
                
                if amount > 0 and is_income(purpose):
                    operations.append({
                        'date': date,
                        'amount': amount,
                        'purpose': purpose[:200],
                        'document': f"{date.strftime('%d.%m.%Y')} {doc_num}"
                    })
                    
            except Exception as e:
                continue
        
        return operations
        
    except Exception as e:
        raise Exception(f"Ошибка парсинга: {e}")


# ========== ПАРСИНГ ВЫПИСКИ ЕНС ==========

def parse_ens_statement(file_path):
    """Парсинг CSV выписки ЕНС"""
    try:
        df = None
        encodings = ['utf-8', 'windows-1251', 'cp1251']
        separators = [';', ',', '\t']
        
        for enc in encodings:
            for sep in separators:
                try:
                    df = pd.read_csv(file_path, sep=sep, encoding=enc, on_bad_lines='skip')
                    if len(df.columns) > 1:
                        break
                except:
                    continue
            if df is not None and len(df.columns) > 1:
                break
        
        if df is None or len(df.columns) <= 1:
            raise Exception("Не удалось определить формат файла")
        
        result = {
            'insurance_accrued': 0.0,
            'insurance_paid': 0.0,
            'insurance_paid_dates': [],
            'penalties': 0.0
        }
        
        df.columns = [str(col).strip().lower() for col in df.columns]
        
        col_operation = None
        col_amount = None
        col_date = None
        col_obligation = None
        col_kbk = None
        
        for col in df.columns:
            if 'наименование операции' in col or 'операция' in col:
                col_operation = col
            elif 'сумма' in col and 'операции' in col:
                col_amount = col
            elif 'дата' in col:
                col_date = col
            elif 'наименование обязательства' in col or 'обязательство' in col:
                col_obligation = col
            elif 'кбк' in col:
                col_kbk = col
        
        if col_amount is None:
            for col in df.columns:
                if df[col].dtype in ['float64', 'int64']:
                    col_amount = col
                    break
        
        if col_operation is None and len(df.columns) > 0:
            col_operation = df.columns[0]
        
        for idx, row in df.iterrows():
            try:
                operation = str(row.get(col_operation, '')).lower() if col_operation else ''
                obligation = str(row.get(col_obligation, '')).lower() if col_obligation else ''
                kbk = str(row.get(col_kbk, '')) if col_kbk else ''
                
                amount = 0.0
                if col_amount:
                    val = row.get(col_amount)
                    if pd.notna(val):
                        amount = safe_float(val)
                
                date_obj = None
                if col_date:
                    date_str = str(row.get(col_date, ''))
                    if date_str and date_str != 'nan':
                        date_obj = parse_date(date_str)
                
                if ('начислено' in operation or 'начисление' in operation) and \
                   ('страховые взносы' in obligation or 'фиксированный размер' in operation):
                    result['insurance_accrued'] += abs(amount)
                
                elif 'пеня' in operation or 'пени' in operation:
                    result['penalties'] += abs(amount)
                
                elif 'уплата' in operation or 'платеж' in operation:
                    if date_obj and date_obj.year == 2026 and amount > 0:
                        result['insurance_paid'] += amount
                        result['insurance_paid_dates'].append(date_obj)
                
                if '18210202000010000160' in kbk and 'уплата' in operation:
                    if date_obj and date_obj.year == 2026:
                        result['insurance_paid'] += amount
                        if date_obj not in result['insurance_paid_dates']:
                            result['insurance_paid_dates'].append(date_obj)
                        
            except Exception as e:
                continue
        
        return result
        
    except Exception as e:
        raise Exception(f"Ошибка парсинга ЕНС: {e}")


# ========== ГЕНЕРАЦИЯ КУДиР ==========

def generate_kudir_from_scratch(operations, output_path, inn=IP_INN, fio=IP_FIO, year=2025):
    """Создание КУДиР с нуля (если шаблон недоступен)"""
    
    wb = Workbook()
    
    # Лист 1 - титульный
    ws = wb.active
    ws.title = "КУДиР"
    
    ws['A1'] = f"Книга учета доходов и расходов ИП {fio}"
    ws['A2'] = f"ИНН {inn}"
    ws['A3'] = f"за {year} год"
    ws['A4'] = "Объект налогообложения: Доходы"
    
    # Таблица доходов
    ws['A6'] = "№ п/п"
    ws['B6'] = "Дата операции"
    ws['C6'] = "Содержание операции"
    ws['D6'] = "Сумма дохода"
    
    sorted_ops = sorted(operations, key=lambda x: x['date'])
    total = 0
    
    for idx, op in enumerate(sorted_ops, 1):
        ws.cell(row=6 + idx, column=1, value=idx)
        ws.cell(row=6 + idx, column=2, value=op['date'].strftime('%d.%m.%Y'))
        ws.cell(row=6 + idx, column=3, value=op['purpose'])
        ws.cell(row=6 + idx, column=4, value=op['amount'])
        total += op['amount']
    
    ws.cell(row=6 + len(sorted_ops) + 1, column=3, value="ИТОГО:")
    ws.cell(row=6 + len(sorted_ops) + 1, column=4, value=total)
    
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 15
    
    wb.save(output_path)
    return total


def fill_kudir_from_template(operations, template_path, output_path, inn=IP_INN, fio=IP_FIO, year=2025):
    """Заполнение шаблона КУДиР данными"""
    
    try:
        wb = load_workbook(template_path)
    except Exception as e:
        # Если шаблон поврежден, создаем с нуля
        return generate_kudir_from_scratch(operations, output_path, inn, fio, year)
    
    sorted_ops = sorted(operations, key=lambda x: x['date'])
    total_income = sum(op['amount'] for op in sorted_ops)
    
    # Пытаемся найти листы
    ws_title = None
    ws_income = None
    
    for sheet in wb.sheetnames:
        if "титул" in sheet.lower() or sheet == "Лист1":
            ws_title = wb[sheet]
        elif "доход" in sheet.lower() or sheet in ["Лист2", "Лист3"]:
            ws_income = wb[sheet]
    
    # Заполняем титульный лист
    if ws_title:
        # Ищем ячейки по их содержимому
        for row in range(1, 50):
            for col in range(1, 50):
                cell = ws_title.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    if "на 20" in cell.value and "год" in cell.value:
                        safe_write(ws_title, row, col, f"{year}")
                    elif "ИНН" in cell.value and "индивидуального" in cell.value.lower():
                        safe_write(ws_title, row, col+1, inn)
                    elif "Объект налогообложения" in cell.value:
                        safe_write(ws_title, row, col+1, "Доходы")
    
    # Заполняем таблицу доходов
    if ws_income:
        # Ищем начало таблицы (строку с "№ п/п")
        start_row = None
        for row in range(1, 100):
            cell = ws_income.cell(row=row, column=1)
            if cell.value and "№ п/п" in str(cell.value):
                start_row = row + 1
                break
        
        if start_row is None:
            start_row = 14
        
        quarterly_totals = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
        
        for idx, op in enumerate(sorted_ops, 1):
            quarter = (op['date'].month - 1) // 3 + 1
            quarterly_totals[quarter] += op['amount']
            
            safe_write(ws_income, start_row + idx - 1, 1, idx)
            safe_write(ws_income, start_row + idx - 1, 2, op['document'])
            safe_write(ws_income, start_row + idx - 1, 3, op['purpose'][:150])
            safe_write(ws_income, start_row + idx - 1, 4, format_currency(op['amount']))
        
        # Итоги
        for row in range(start_row, start_row + len(sorted_ops) + 20):
            cell = ws_income.cell(row=row, column=1)
            if cell.value and isinstance(cell.value, str):
                if "Итого за I квартал" in cell.value:
                    safe_write(ws_income, row, 4, format_currency(quarterly_totals[1]))
                elif "Итого за II квартал" in cell.value:
                    safe_write(ws_income, row, 4, format_currency(quarterly_totals[2]))
                elif "Итого за III квартал" in cell.value:
                    safe_write(ws_income, row, 4, format_currency(quarterly_totals[3]))
                elif "Итого за IV квартал" in cell.value:
                    safe_write(ws_income, row, 4, format_currency(quarterly_totals[4]))
                elif "Итого за полугодие" in cell.value:
                    safe_write(ws_income, row, 4, format_currency(quarterly_totals[1] + quarterly_totals[2]))
                elif "Итого за 9 месяцев" in cell.value:
                    safe_write(ws_income, row, 4, format_currency(quarterly_totals[1] + quarterly_totals[2] + quarterly_totals[3]))
                elif "Итого за год" in cell.value:
                    safe_write(ws_income, row, 4, format_currency(total_income))
                elif cell.value.strip() == "010":
                    safe_write(ws_income, row, 4, format_currency(total_income))
                elif cell.value.strip() == "020":
                    safe_write(ws_income, row, 4, 0)
                elif cell.value.strip() == "040":
                    safe_write(ws_income, row, 4, format_currency(total_income))
    
    wb.save(output_path)
    return total_income


# ========== ГЕНЕРАЦИЯ ДЕКЛАРАЦИИ ==========

def generate_declaration_from_scratch(operations, ens_data, output_excel, output_xml, inn=IP_INN, fio=IP_FIO):
    """Создание декларации с нуля"""
    
    quarterly = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
    for op in operations:
        quarter = (op['date'].month - 1) // 3 + 1
        quarterly[quarter] += op['amount']
    
    total_income = sum(quarterly.values())
    tax_rate = 6
    tax_amount = total_income * tax_rate / 100
    
    paid_in_2025 = any(d.year == 2025 for d in ens_data.get('insurance_paid_dates', []))
    insurance_paid = ens_data.get('insurance_paid', 0)
    
    if paid_in_2025:
        tax_payable = max(0, tax_amount - insurance_paid)
        deductible = insurance_paid
    else:
        tax_payable = tax_amount
        deductible = 0
    
    cum_income = {
        1: quarterly[1],
        2: quarterly[1] + quarterly[2],
        3: quarterly[1] + quarterly[2] + quarterly[3],
        4: total_income
    }
    
    cum_tax = {
        1: cum_income[1] * tax_rate / 100,
        2: cum_income[2] * tax_rate / 100,
        3: cum_income[3] * tax_rate / 100,
        4: tax_amount
    }
    
    if paid_in_2025:
        cum_deductible = {
            1: min(cum_tax[1], deductible),
            2: min(cum_tax[2], deductible),
            3: min(cum_tax[3], deductible),
            4: min(cum_tax[4], deductible)
        }
    else:
        cum_deductible = {1: 0, 2: 0, 3: 0, 4: 0}
    
    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Декларация УСН"
    
    ws['A1'] = "Налоговая декларация по УСН"
    ws['A2'] = f"ИП {fio}"
    ws['A3'] = f"ИНН {inn}"
    ws['A4'] = "за 2025 год"
    
    ws['A6'] = "Раздел 2.1.1. Доходы"
    ws['A8'] = "Доход за 1 квартал"
    ws['B8'] = cum_income[1]
    ws['A9'] = "Доход за полугодие"
    ws['B9'] = cum_income[2]
    ws['A10'] = "Доход за 9 месяцев"
    ws['B10'] = cum_income[3]
    ws['A11'] = "Доход за год"
    ws['B11'] = cum_income[4]
    ws['A12'] = "Налоговая ставка (%)"
    ws['B12'] = tax_rate
    ws['A13'] = "Сумма налога за год"
    ws['B13'] = tax_amount
    ws['A14'] = "Сумма страховых взносов"
    ws['B14'] = cum_deductible[4]
    
    ws['A16'] = "Раздел 1.1. Налог к уплате"
    ws['A17'] = "ОКТМО"
    ws['B17'] = IP_OKTMO
    ws['A18'] = "Налог к уплате за год"
    ws['B18'] = tax_payable
    
    wb.save(output_excel)
    
    # XML
    fio_parts = fio.split()
    last_name = fio_parts[0] if len(fio_parts) > 0 else ""
    first_name = fio_parts[1] if len(fio_parts) > 1 else ""
    patronymic = fio_parts[2] if len(fio_parts) > 2 else ""
    
    xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<Файл xmlns="urn:ФНС-СХД-Декл-УСН-2025-1">
    <Документ>
        <КНД>1152017</КНД>
        <ДатаДок>{datetime.now().strftime('%Y-%m-%d')}</ДатаДок>
        <НомКорр>0</НомКорр>
    </Документ>
    <НалогПериод>
        <НомерПериода>34</НомерПериода>
        <ОтчетныйГод>2025</ОтчетныйГод>
    </НалогПериод>
    <Налогоплательщик>
        <ИНН>{inn}</ИНН>
        <ИП>
            <ФИО>
                <Фамилия>{last_name}</Фамилия>
                <Имя>{first_name}</Имя>
                <Отчество>{patronymic}</Отчество>
            </ФИО>
        </ИП>
    </Налогоплательщик>
    <Показатели>
        <Раздел1_1>
            <ОКТМО>{IP_OKTMO}</ОКТМО>
            <СумНал100>{int(tax_payable)}</СумНал100>
        </Раздел1_1>
        <Раздел2_1_1>
            <СумДоход110>{int(cum_income[1])}</СумДоход110>
            <СумДоход111>{int(cum_income[2])}</СумДоход111>
            <СумДоход112>{int(cum_income[3])}</СумДоход112>
            <СумДоход113>{int(cum_income[4])}</СумДоход113>
            <НалСтавка120>{tax_rate}</НалСтавка120>
            <СумИсчисНал133>{int(tax_amount)}</СумИсчисНал133>
            <СумУплНал143>{int(cum_deductible[4])}</СумУплНал143>
        </Раздел2_1_1>
    </Показатели>
</Файл>'''
    
    with open(output_xml, 'w', encoding='utf-8') as f:
        f.write(xml_content)
    
    return tax_payable, total_income


def fill_declaration_from_template(operations, ens_data, template_path, output_excel, output_xml, 
                                    inn=IP_INN, fio=IP_FIO, okved=IP_OKVED, phone=IP_PHONE):
    """Заполнение шаблона декларации данными"""
    
    try:
        wb = load_workbook(template_path)
    except Exception:
        return generate_declaration_from_scratch(operations, ens_data, output_excel, output_xml, inn, fio)
    
    # Расчет доходов
    quarterly = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
    for op in operations:
        quarter = (op['date'].month - 1) // 3 + 1
        quarterly[quarter] += op['amount']
    
    total_income = sum(quarterly.values())
    tax_rate = 6
    tax_amount = total_income * tax_rate / 100
    
    paid_in_2025 = any(d.year == 2025 for d in ens_data.get('insurance_paid_dates', []))
    insurance_paid = ens_data.get('insurance_paid', 0)
    
    if paid_in_2025:
        tax_payable = max(0, tax_amount - insurance_paid)
        deductible = insurance_paid
    else:
        tax_payable = tax_amount
        deductible = 0
    
    cum_income = {
        1: quarterly[1],
        2: quarterly[1] + quarterly[2],
        3: quarterly[1] + quarterly[2] + quarterly[3],
        4: total_income
    }
    
    cum_tax = {
        1: cum_income[1] * tax_rate / 100,
        2: cum_income[2] * tax_rate / 100,
        3: cum_income[3] * tax_rate / 100,
        4: tax_amount
    }
    
    if paid_in_2025:
        cum_deductible = {
            1: min(cum_tax[1], deductible),
            2: min(cum_tax[2], deductible),
            3: min(cum_tax[3], deductible),
            4: min(cum_tax[4], deductible)
        }
    else:
        cum_deductible = {1: 0, 2: 0, 3: 0, 4: 0}
    
    ws1 = wb["стр.1"] if "стр.1" in wb.sheetnames else wb.active
    
    # Заполнение
    for r in range(1, 200):
        code_cell = ws1.cell(row=r, column=3).value
        if code_cell:
            code = str(code_cell).strip()
            if code == "010":
                safe_write(ws1, r, 4, format_currency(cum_income[1]))
            elif code == "011":
                safe_write(ws1, r, 4, format_currency(cum_income[2]))
            elif code == "012":
                safe_write(ws1, r, 4, format_currency(cum_income[3]))
            elif code == "013":
                safe_write(ws1, r, 4, format_currency(cum_income[4]))
            elif code == "020":
                safe_write(ws1, r, 4, tax_rate)
            elif code == "030":
                safe_write(ws1, r, 4, format_currency(cum_tax[1]))
            elif code == "031":
                safe_write(ws1, r, 4, format_currency(cum_tax[2]))
            elif code == "032":
                safe_write(ws1, r, 4, format_currency(cum_tax[3]))
            elif code == "033":
                safe_write(ws1, r, 4, format_currency(cum_tax[4]))
            elif code == "040":
                safe_write(ws1, r, 4, format_currency(cum_deductible[1]))
            elif code == "041":
                safe_write(ws1, r, 4, format_currency(cum_deductible[2]))
            elif code == "042":
                safe_write(ws1, r, 4, format_currency(cum_deductible[3]))
            elif code == "043":
                safe_write(ws1, r, 4, format_currency(cum_deductible[4]))
            elif code == "050":
                safe_write(ws1, r, 4, IP_OKTMO)
            elif code == "060":
                safe_write(ws1, r, 4, format_currency(tax_payable))
    
    wb.save(output_excel)
    
    # XML
    fio_parts = fio.split()
    last_name = fio_parts[0] if len(fio_parts) > 0 else ""
    first_name = fio_parts[1] if len(fio_parts) > 1 else ""
    patronymic = fio_parts[2] if len(fio_parts) > 2 else ""
    
    xml_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<Файл xmlns="urn:ФНС-СХД-Декл-УСН-2025-1">
    <Документ>
        <КНД>1152017</КНД>
        <ДатаДок>{datetime.now().strftime('%Y-%m-%d')}</ДатаДок>
        <НомКорр>0</НомКорр>
    </Документ>
    <НалогПериод>
        <НомерПериода>34</НомерПериода>
        <ОтчетныйГод>2025</ОтчетныйГод>
    </НалогПериод>
    <Налогоплательщик>
        <ИНН>{inn}</ИНН>
        <ИП>
            <ФИО>
                <Фамилия>{last_name}</Фамилия>
                <Имя>{first_name}</Имя>
                <Отчество>{patronymic}</Отчество>
            </ФИО>
        </ИП>
    </Налогоплательщик>
    <Показатели>
        <Раздел1_1>
            <ОКТМО>{IP_OKTMO}</ОКТМО>
            <СумНал100>{int(tax_payable)}</СумНал100>
        </Раздел1_1>
        <Раздел2_1_1>
            <СумДоход110>{int(cum_income[1])}</СумДоход110>
            <СумДоход111>{int(cum_income[2])}</СумДоход111>
            <СумДоход112>{int(cum_income[3])}</СумДоход112>
            <СумДоход113>{int(cum_income[4])}</СумДоход113>
            <НалСтавка120>{tax_rate}</НалСтавка120>
            <СумИсчисНал133>{int(tax_amount)}</СумИсчисНал133>
            <СумУплНал143>{int(cum_deductible[4])}</СумУплНал143>
        </Раздел2_1_1>
    </Показатели>
</Файл>'''
    
    with open(output_xml, 'w', encoding='utf-8') as f:
        f.write(xml_content)
    
    return tax_payable, total_income


# ========== ОБРАБОТЧИКИ TELEGRAM ==========

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_sessions[user_id] = UserSession(user_id)
    
    await update.message.reply_text(
        "🤖 *Бот для подготовки отчетности ИП на УСН*\n\n"
        "Я помогу вам:\n"
        "📊 Сформировать КУДиР\n"
        "📝 Заполнить декларацию по УСН\n"
        "💰 Рассчитать налог к уплате\n\n"
        "*Как работать:*\n"
        "1️⃣ Загрузите выписки с расчетных счетов (Excel)\n"
        "2️⃣ Загрузите выписку с ЕНС (CSV)\n"
        "3️⃣ Введите /report\n\n"
        "📌 *Сроки за 2025 год:*\n"
        "• Декларацию сдать до *27 апреля 2026*\n"
        "• Налог уплатить до *28 апреля 2026*\n\n"
        "⚠️ Если не сдать декларацию в срок — налоговая может заблокировать счет.\n"
        "Просрочка уплаты налога счет не блокирует — только пени.",
        parse_mode="Markdown"
    )


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    if user_id not in user_sessions:
        user_sessions[user_id] = UserSession(user_id)
    
    session = user_sessions[user_id]
    document = update.message.document
    filename = document.file_name.lower()
    
    file = await context.bot.get_file(document.file_id)
    
    with tempfile.NamedTemporaryFile(suffix=os.path.splitext(filename)[1], delete=False) as tmp:
        await file.download_to_drive(tmp.name)
        tmp_path = tmp.name
    
    try:
        if filename.endswith(('.xlsx', '.xls')):
            await update.message.reply_text("📥 Обрабатываю выписку из банка...")
            operations = parse_bank_statement(tmp_path)
            
            if operations:
                session.add_bank_operations(operations)
                total = sum(op['amount'] for op in operations)
                await update.message.reply_text(
                    f"✅ Найдено {len(operations)} операций\n"
                    f"💰 Сумма доходов: {total:,.2f} ₽\n\n"
                    f"Загружайте другие выписки или пришлите выписку с ЕНС (CSV)."
                )
            else:
                await update.message.reply_text(
                    "⚠️ В выписке не найдено доходов.\n"
                    "Проверьте, что в файле есть поступления от покупателей."
                )
        
        elif filename.endswith('.csv'):
            await update.message.reply_text("📥 Обрабатываю выписку с ЕНС...")
            ens_data = parse_ens_statement(tmp_path)
            session.set_ens_data(ens_data)
            
            paid_in_2025 = any(d.year == 2025 for d in ens_data['insurance_paid_dates'])
            
            await update.message.reply_text(
                f"✅ Выписка ЕНС обработана!\n\n"
                f"📌 Страховые взносы:\n"
                f"• Начислено: {ens_data['insurance_accrued']:,.2f} ₽\n"
                f"• Уплачено: {ens_data['insurance_paid']:,.2f} ₽\n"
                f"• Уплачено в 2025: {'Да' if paid_in_2025 else 'Нет'}\n"
                f"• Пени: {ens_data['penalties']:,.2f} ₽\n\n"
                f"Теперь введите /report"
            )
        
        else:
            await update.message.reply_text("❌ Поддерживаются только .xlsx, .xls и .csv")
    
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")
    
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


async def report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    if user_id not in user_sessions:
        await update.message.reply_text("Сначала загрузите выписки (/start)")
        return
    
    session = user_sessions[user_id]
    
    if not session.bank_operations:
        await update.message.reply_text("⚠️ Сначала загрузите выписки с расчетных счетов")
        return
    
    if not session.ens_data.get('insurance_accrued') and not session.ens_data.get('insurance_paid'):
        await update.message.reply_text("⚠️ Сначала загрузите выписку с ЕНС")
        return
    
    await update.message.reply_text("🔄 Формирую отчетность... Это может занять несколько секунд.")
    
    try:
        # Сортируем все операции
        all_ops = []
        for ops in session.bank_operations:
            all_ops.extend(ops)
        all_ops.sort(key=lambda x: x['date'])
        
        # Пытаемся использовать шаблоны
        kudir_template = os.path.join(TEMPLATES_DIR, "KUDIR_template.xlsx")
        decl_template = os.path.join(TEMPLATES_DIR, "Declaration_template.xlsx")
        
        # КУДиР
        kudir_path = os.path.join(OUTPUT_DIR, f"kudir_{user_id}.xlsx")
        if os.path.exists(kudir_template):
            total_income = fill_kudir_from_template(
                all_ops, kudir_template, kudir_path,
                inn=IP_INN, fio=IP_FIO, year=2025
            )
        else:
            total_income = generate_kudir_from_scratch(
                all_ops, kudir_path, inn=IP_INN, fio=IP_FIO, year=2025
            )
        
        # Декларация
        decl_excel = os.path.join(OUTPUT_DIR, f"declaration_{user_id}.xlsx")
        decl_xml = os.path.join(OUTPUT_DIR, f"declaration_{user_id}.xml")
        
        if os.path.exists(decl_template):
            tax_payable, total_income = fill_declaration_from_template(
                all_ops, session.ens_data, decl_template, decl_excel, decl_xml,
                inn=IP_INN, fio=IP_FIO
            )
        else:
            tax_payable, total_income = generate_declaration_from_scratch(
                all_ops, session.ens_data, decl_excel, decl_xml,
                inn=IP_INN, fio=IP_FIO
            )
        
        await update.message.reply_text(
            f"✅ *Отчетность готова!*\n\n"
            f"📊 *Доход за 2025:* {total_income:,.2f} ₽\n"
            f"💰 *Налог к уплате:* {tax_payable:,.2f} ₽\n\n"
            f"📌 *Сроки:*\n"
            f"• Декларацию сдать до *27 апреля 2026*\n"
            f"• Налог уплатить до *28 апреля 2026*\n\n"
            f"⚠️ Если не сдать декларацию в срок — заблокируют счет.\n"
            f"Просрочка уплаты налога счет не блокирует.\n\n"
            f"📎 Отправляю файлы...",
            parse_mode="Markdown"
        )
        
        with open(kudir_path, 'rb') as f:
            await update.message.reply_document(f, filename="КУДиР_2025.xlsx", caption="📘 Книга учета доходов и расходов")
        
        with open(decl_excel, 'rb') as f:
            await update.message.reply_document(f, filename="Декларация_УСН_2025.xlsx", caption="📝 Декларация по УСН (Excel для проверки)")
        
        with open(decl_xml, 'rb') as f:
            await update.message.reply_document(f, filename="declaration_usn_2025.xml", caption="📎 XML для загрузки в ЛК ФНС")
        
        await update.message.reply_text(
            "🎉 *Готово!*\n\n"
            "Что дальше:\n"
            "1. Проверьте декларацию в Excel\n"
            "2. Загрузите XML в Личный кабинет ИП на сайте ФНС\n"
            "3. Подпишите электронной подписью и отправьте\n"
            "4. Уплатите налог до 28 апреля 2026\n\n"
            "💡 *Важно:* взносы, уплаченные в 2026 году, не уменьшают налог за 2025."
        )
    
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")
        import traceback
        traceback.print_exc()


async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id in user_sessions:
        user_sessions[user_id].reset()
        await update.message.reply_text("🔄 Данные сброшены. Начните с /start")
    else:
        await update.message.reply_text("Нет активной сессии. Используйте /start")


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🤖 *Помощь*\n\n"
        "*Команды:*\n"
        "/start — начать работу\n"
        "/report — сформировать отчетность\n"
        "/reset — сбросить все данные\n"
        "/help — эта справка\n\n"
        "*Файлы:*\n"
        "• Сначала загрузите Excel-выписки из банков\n"
        "• Затем загрузите CSV-выписку с ЕНС\n"
        "• Введите /report\n\n"
        "*Сроки за 2025 год:*\n"
        "• Декларация: до 27 апреля 2026\n"
        "• Уплата налога: до 28 апреля 2026\n\n"
        "*Важно:*\n"
        "• За несдачу декларации — блокировка счета\n"
        "• За просрочку уплаты налога — только пени",
        parse_mode="Markdown"
    )


def main():
    if not BOT_TOKEN:
        print("❌ Ошибка: BOT_TOKEN не задан в .env файле")
        sys.exit(1)
    
    app = Application.builder().token(BOT_TOKEN).build()
    
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("report", report))
    app.add_handler(CommandHandler("reset", reset))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    
    print("🤖 Бот запущен...")
    print(f"📁 Папка с выгрузкой: {OUTPUT_DIR}")
    
    app.run_polling()


if __name__ == "__main__":
    main()